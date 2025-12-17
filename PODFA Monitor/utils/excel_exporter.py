"""
PBS 2.0 Enhanced Excel Exporter
================================

향상된 엑셀 저장 시스템
- 현재 보이는 영역만 저장
- 고품질 차트 이미지 삽입
- 자동 차트 생성
- 메타데이터 포함
"""

import os
import io
import time
import logging
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image as PILImage

from core.data_processor import DataPoint
from ui.chart_widget import ChartWidget


@dataclass
class ExportOptions:
    """엑셀 내보내기 옵션"""
    export_type: str = "viewport"     # "viewport" | "all" | "range"
    include_chart_image: bool = True  # 차트 이미지 포함
    include_excel_chart: bool = True  # 엑셀 네이티브 차트 포함
    include_metadata: bool = True     # 메타데이터 포함
    include_statistics: bool = True   # 통계 정보 포함
    
    # 이미지 옵션
    image_format: str = "png"        # png | jpg | svg
    image_width: int = 1800          # 픽셀 (더 넓은 차트를 위해 증가)
    image_height: int = 1800         # 픽셀 (높이 3배 증가)
    image_dpi: int = 450            # DPI (고품질 이미지를 위해 증가)
    
    # 데이터 옵션
    use_calibrated_values: bool = True  # 캘리브레이션된 값 사용
    include_quality_scores: bool = False # 품질 점수 포함
    decimal_places: int = 3           # 소수점 자릿수
    
    # 스타일 옵션
    apply_formatting: bool = True     # 서식 적용
    freeze_header: bool = True        # 헤더 고정
    auto_column_width: bool = True    # 자동 열 너비


@dataclass
class ExportResult:
    """내보내기 결과"""
    success: bool
    file_path: str
    data_points: int
    sheets_created: List[str]
    file_size: int
    export_time: float
    error_message: Optional[str] = None


class EnhancedExcelExporter:
    """
    💾 고급 엑셀 내보내기 시스템
    
    Features:
    - 현재 뷰포트 데이터 저장
    - 고품질 차트 이미지 삽입
    - 엑셀 네이티브 차트 생성
    - 메타데이터 및 통계 포함
    - 전문적인 서식 적용
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self._init_styles()
    
    def _init_styles(self):
        """스타일 초기화"""
        # 헤더 스타일
        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        self.header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_style.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"), 
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")
        
        # 데이터 스타일
        self.data_style = NamedStyle(name="data")
        self.data_style.font = Font(name='Calibri', size=10)
        self.data_style.border = Border(
            left=Side(border_style="thin", color="CCCCCC"),
            right=Side(border_style="thin", color="CCCCCC"),
            top=Side(border_style="thin", color="CCCCCC"),
            bottom=Side(border_style="thin", color="CCCCCC")
        )
        self.data_style.alignment = Alignment(horizontal="center", vertical="center")
    
    def export_workbench(
        self,
        workbench_name: str,
        chart_widgets: List[ChartWidget],
        file_path: str,
        options: Optional[ExportOptions] = None
    ) -> ExportResult:
        """
        워크벤치 전체 내보내기
        
        Args:
            workbench_name: 워크벤치 이름
            chart_widgets: 차트 위젯 리스트
            file_path: 저장할 파일 경로
            options: 내보내기 옵션
            
        Returns:
            내보내기 결과
        """
        start_time = time.time()
        options = options or ExportOptions()
        
        try:
            # 워크북 생성
            wb = Workbook()
            wb.remove(wb.active)  # 기본 시트 제거
            
            # 스타일 등록
            if options.apply_formatting:
                wb.add_named_style(self.header_style)
                wb.add_named_style(self.data_style)
            
            sheets_created = []
            total_data_points = 0
            
            # 요약 시트 생성
            if len(chart_widgets) > 1:
                summary_ws = wb.create_sheet("Summary")
                self._create_summary_sheet(summary_ws, workbench_name, chart_widgets, options)
                sheets_created.append("Summary")
            
            # 각 차트 위젯별 시트 생성
            for i, widget in enumerate(chart_widgets):
                data_points = self._export_chart_widget(wb, widget, f"Chart_{i+1}", options)
                sheets_created.extend([f"Chart_{i+1}_Data", f"Chart_{i+1}_Chart"])
                total_data_points += data_points
            
            # 메타데이터 시트 생성
            if options.include_metadata:
                meta_ws = wb.create_sheet("Metadata")
                self._create_metadata_sheet(meta_ws, workbench_name, chart_widgets, options)
                sheets_created.append("Metadata")
            
            # 파일 저장
            wb.save(file_path)
            
            # 결과 반환
            export_time = time.time() - start_time
            file_size = os.path.getsize(file_path)
            
            self.logger.info(f"워크벤치 내보내기 완료: {file_path} ({file_size:,} bytes)")
            
            return ExportResult(
                success=True,
                file_path=file_path,
                data_points=total_data_points,
                sheets_created=sheets_created,
                file_size=file_size,
                export_time=export_time
            )
            
        except Exception as e:
            export_time = time.time() - start_time
            self.logger.error(f"워크벤치 내보내기 실패: {e}")
            
            return ExportResult(
                success=False,
                file_path=file_path,
                data_points=0,
                sheets_created=[],
                file_size=0,
                export_time=export_time,
                error_message=str(e)
            )
    
    def _export_chart_widget(
        self,
        wb: Workbook,
        widget: ChartWidget,
        sheet_name: str,
        options: ExportOptions
    ) -> int:
        """차트 위젯 내보내기"""
        # 데이터 추출
        if options.export_type == "viewport":
            data_points = widget.get_visible_data()
        else:  # "all"
            data_points = widget.data_points
        
        if not data_points:
            self.logger.warning(f"차트 '{widget.name}'에 데이터가 없습니다")
            return 0
        
        # 통합 시트 생성 (이미지 + 데이터)
        combined_ws = wb.create_sheet(sheet_name)
        self._create_combined_sheet(combined_ws, widget, data_points, options)
        
        return len(data_points)
    
    def _create_combined_sheet(
        self,
        worksheet: Worksheet,
        widget: ChartWidget,
        data_points: List[DataPoint],
        options: ExportOptions
    ):
        """통합 시트 생성 (이미지 + 데이터)"""
        current_row = 1
        
        # 제목
        title_cell = worksheet.cell(row=current_row, column=1, value=f"Chart: {widget.name}")
        if options.apply_formatting:
            title_cell.font = Font(size=16, bold=True, color="2F5597")
        current_row += 2
        
        # 차트 이미지 삽입
        if options.include_chart_image:
            try:
                # 차트 이미지 생성
                image_bytes = widget.generate_image(
                    format=options.image_format,
                    width=options.image_width,
                    height=options.image_height,
                    dpi=options.image_dpi
                )
                
                if image_bytes is not None:
                    try:
                        # 메모리에서 직접 이미지 생성
                        img_stream = io.BytesIO(image_bytes)
                        img = Image(img_stream)
                        
                        # 적절한 크기로 조정 (더 큰 이미지를 위해 최대 크기 증가)
                        img.width = options.image_width * 0.75
                        img.height = options.image_height * 0.75
                        
                        # Excel에 이미지 삽입
                        worksheet.add_image(img, f"A{current_row}")
                        
                        # 이미지 아래 여백 추가 (대략 이미지 높이의 픽셀을 행으로 변환)
                        current_row += max(25, int(img.height / 15))  # 대략 15픽셀 = 1행
                        
                        self.logger.info(f"차트 '{widget.name}' 이미지 삽입 완료")
                        
                    except Exception as img_error:
                        self.logger.error(f"이미지 삽입 실패 ({widget.name}): {img_error}")
                        current_row += 5  # 이미지 실패 시 기본 여백
                else:
                    self.logger.warning(f"차트 '{widget.name}' 이미지 생성 실패")
                    current_row += 5
                    
            except Exception as e:
                self.logger.error(f"차트 이미지 처리 중 오류 ({widget.name}): {e}")
                current_row += 5
        
        # 데이터 섹션 제목
        data_title_cell = worksheet.cell(row=current_row, column=1, value="📊 Data")
        if options.apply_formatting:
            data_title_cell.font = Font(size=14, bold=True, color="2F5597")
        current_row += 2
        
        # 데이터 헤더 생성
        headers = ["Time", "Raw Value"]
        
        if any(dp.filtered_value is not None for dp in data_points):
            headers.append("Filtered Value")
            
        if options.use_calibrated_values and any(dp.calibrated_value is not None for dp in data_points):
            headers.append("Weight (g)")
            
        if options.include_quality_scores:
            headers.append("Quality Score")
        
        # 헤더 작성
        header_row = current_row
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col, value=header)
            if options.apply_formatting:
                cell.style = "header"
        
        current_row += 1
        
        # 데이터 작성
        for dp in data_points:
            col = 1
            
            # 시간 (상대적)
            time_diff = dp.timestamp - data_points[0].timestamp if data_points else 0
            worksheet.cell(row=current_row, column=col, value=f"{time_diff:.1f}s")
            col += 1
            
            # 원시값
            raw_val = round(dp.raw_value, options.decimal_places)
            worksheet.cell(row=current_row, column=col, value=raw_val)
            col += 1
            
            # 필터링된 값
            if any(dp.filtered_value is not None for dp in data_points):
                if dp.filtered_value is not None:
                    filtered_val = round(dp.filtered_value, options.decimal_places)
                    worksheet.cell(row=current_row, column=col, value=filtered_val)
                col += 1
            
            # 캘리브레이션된 값
            if options.use_calibrated_values and any(dp.calibrated_value is not None for dp in data_points):
                if dp.calibrated_value is not None:
                    cal_val = round(dp.calibrated_value, options.decimal_places)
                    worksheet.cell(row=current_row, column=col, value=cal_val)
                col += 1
            
            # 품질 점수
            if options.include_quality_scores:
                worksheet.cell(row=current_row, column=col, value=getattr(dp, 'quality_score', 1.0))
                col += 1
            
            current_row += 1
        
        # 통계 테이블 추가 (39행부터)
        if options.include_statistics and data_points:
            stats_row = 39  # 39행부터 시작
            
            # 통계 테이블 제목
            stats_title_cell = worksheet.cell(row=stats_row, column=1, value="📈 Statistics")
            if options.apply_formatting:
                stats_title_cell.font = Font(size=14, bold=True, color="2F5597")
            stats_row += 2
            
            # 통계 데이터 계산
            values = [
                dp.calibrated_value if dp.calibrated_value is not None 
                else dp.filtered_value if dp.filtered_value is not None
                else dp.raw_value
                for dp in data_points
            ]
            
            # 통계 헤더
            stats_headers = ["Metric", "Value"]
            for col, header in enumerate(stats_headers, 1):
                cell = worksheet.cell(row=stats_row, column=col, value=header)
                if options.apply_formatting:
                    cell.style = "header"
            stats_row += 1
            
            # 통계 값 계산 및 표시
            import numpy as np
            stats_data = [
                ("Count", len(values)),
                ("Max", f"{max(values):.3f}"),
                ("Min", f"{min(values):.3f}"),
                ("Mean", f"{np.mean(values):.3f}"),
                ("Std Dev", f"{np.std(values):.3f}"),
                ("Median", f"{np.median(values):.3f}")
            ]
            
            # Max Sensor Value from widget if available
            if hasattr(widget, 'max_sensor_value') and widget.max_sensor_value != float('-inf'):
                stats_data.insert(1, ("Peak Value", f"{widget.max_sensor_value:.3f}"))
            
            for metric, value in stats_data:
                worksheet.cell(row=stats_row, column=1, value=metric)
                worksheet.cell(row=stats_row, column=2, value=value)
                if options.apply_formatting:
                    for col in range(1, 3):
                        worksheet.cell(row=stats_row, column=col).style = "data"
                stats_row += 1
        
        # 열 너비 자동 조정
        if options.auto_column_width:
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # 헤더 행 고정
        if options.freeze_header and len(data_points) > 0:
            # 데이터 헤더 바로 다음 행에서 고정 (데이터 첫 행)
            freeze_row = header_row + 1
            worksheet.freeze_panes = f"A{freeze_row}"
    
    def _create_summary_sheet(
        self,
        worksheet: Worksheet,
        workbench_name: str,
        chart_widgets: List[ChartWidget],
        options: ExportOptions
    ):
        """요약 시트 생성"""
        # 제목
        title_cell = worksheet.cell(row=1, column=1, value=f"Workbench Summary: {workbench_name}")
        if options.apply_formatting:
            title_cell.font = Font(size=16, bold=True)
        
        # 헤더
        headers = ["Chart Name", "Data Points", "Max Value", "Mean Value", "Export Type"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col, value=header)
            if options.apply_formatting:
                cell.style = "header"
        
        # 각 차트 정보
        for row, widget in enumerate(chart_widgets, 4):
            if options.export_type == "viewport":
                data_points = widget.get_visible_data()
            else:
                data_points = widget.data_points
            
            if data_points:
                values = [
                    dp.calibrated_value or dp.filtered_value or dp.raw_value
                    for dp in data_points
                ]
                max_val = max(values)
                mean_val = np.mean(values)
            else:
                max_val = 0
                mean_val = 0
            
            worksheet.cell(row=row, column=1, value=widget.name)
            worksheet.cell(row=row, column=2, value=len(data_points))
            worksheet.cell(row=row, column=3, value=round(max_val, options.decimal_places))
            worksheet.cell(row=row, column=4, value=round(mean_val, options.decimal_places))
            worksheet.cell(row=row, column=5, value=options.export_type.title())
            
            # 스타일 적용
            if options.apply_formatting:
                for col in range(1, 6):
                    worksheet.cell(row=row, column=col).style = "data"
        
        # 열 너비 조정
        if options.auto_column_width:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 25)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_metadata_sheet(
        self,
        worksheet: Worksheet,
        workbench_name: str,
        chart_widgets: List[ChartWidget],
        options: ExportOptions
    ):
        """메타데이터 시트 생성"""
        row = 1
        
        # 기본 정보
        worksheet.cell(row=row, column=1, value="Export Information")
        if options.apply_formatting:
            worksheet.cell(row=row, column=1).font = Font(size=14, bold=True)
        row += 2
        
        info_data = [
            ("Workbench Name", workbench_name),
            ("Export Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Export Type", options.export_type.title()),
            ("Charts Count", len(chart_widgets)),
            ("Include Chart Image", "Yes" if options.include_chart_image else "No"),
            ("Include Excel Chart", "Yes" if options.include_excel_chart else "No"),
            ("Use Calibrated Values", "Yes" if options.use_calibrated_values else "No"),
            ("Decimal Places", options.decimal_places),
            ("Image Format", options.image_format.upper()),
            ("Image Resolution", f"{options.image_width}x{options.image_height} @ {options.image_dpi} DPI")
        ]
        
        for key, value in info_data:
            worksheet.cell(row=row, column=1, value=key)
            worksheet.cell(row=row, column=2, value=str(value))
            if options.apply_formatting:
                worksheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1
        
        # 차트별 상세 정보
        if len(chart_widgets) > 1:
            row += 2
            worksheet.cell(row=row, column=1, value="Chart Details")
            if options.apply_formatting:
                worksheet.cell(row=row, column=1).font = Font(size=14, bold=True)
            row += 1
            
            for i, widget in enumerate(chart_widgets, 1):
                worksheet.cell(row=row, column=1, value=f"Chart {i}: {widget.name}")
                if options.apply_formatting:
                    worksheet.cell(row=row, column=1).font = Font(bold=True)
                row += 1
                
                if options.export_type == "viewport":
                    data_points = widget.get_visible_data()
                else:
                    data_points = widget.data_points
                
                chart_info = [
                    ("  Data Points", len(data_points)),
                    ("  Max Sensor Value", f"{widget.max_sensor_value:.3f}g"),
                    ("  Chart Style", "Standard"),
                    ("  Auto Scroll", "Yes")
                ]
                
                for key, value in chart_info:
                    worksheet.cell(row=row, column=1, value=key)
                    worksheet.cell(row=row, column=2, value=str(value))
                    row += 1
                
                row += 1
        
        # 열 너비 조정
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 30
    
    def _create_single_chart_metadata(
        self,
        worksheet: Worksheet,
        widget: ChartWidget,
        options: ExportOptions
    ):
        """단일 차트 메타데이터 생성"""
        self._create_metadata_sheet(worksheet, widget.name, [widget], options)


# 호환성을 위한 별명
ExcelExporter = EnhancedExcelExporter